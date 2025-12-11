# /// script
# requires-python = ">=3.10"
# dependencies = [
#     "openpyxl",
# ]
# ///
"""
Extract questions from Excel file and save as JSON.
Usage: uv run extract_questions.py
"""

import json
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).parent
ROOT_DIR = SCRIPT_DIR.parent
EXCEL_PATH = ROOT_DIR / "temp" / "questions.xlsx"
OUTPUT_PATH = ROOT_DIR / "analysis" / "questions.json"


def find_question_column(headers: list[str]) -> int | None:
    """Find column index containing 'question' (case-insensitive)."""
    for idx, header in enumerate(headers):
        if header and "question" in str(header).lower():
            return idx
    return None


def extract_questions() -> list[str]:
    """Extract questions from Excel file."""
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(
            f"Excel file not found at {EXCEL_PATH}\n"
            "Please download the Google Sheet as Excel and place it there."
        )

    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file is empty")

    headers = list(rows[0])
    question_col = find_question_column(headers)

    if question_col is None:
        raise ValueError(
            f"Could not find a column with 'question' in header.\n"
            f"Available headers: {headers}"
        )

    print(f"Found question column: '{headers[question_col]}' (index {question_col})")

    questions = []
    for row in rows[1:]:
        if question_col < len(row) and row[question_col]:
            q = str(row[question_col]).strip()
            if q:
                questions.append(q)

    return questions


def main():
    print(f"Reading Excel from: {EXCEL_PATH}")
    questions = extract_questions()
    print(f"Extracted {len(questions)} questions")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(questions, f, indent=2)

    print(f"Saved questions to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
